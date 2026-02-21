import os, re, ast
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

BASE = r"c:\Users\C203-PROF\Desktop\paldez\CAV-CCS-DCSN03C_2-2526-2nd-Practical Exam Submission-IT 103-78273"

def find_root(path):
    if os.path.exists(os.path.join(path, "app.py")):
        return path
    for item in os.listdir(path):
        sub = os.path.join(path, item)
        if os.path.isdir(sub) and item not in ('.', 'venv', '__pycache__'):
            if os.path.exists(os.path.join(sub, "app.py")):
                return sub
            for item2 in os.listdir(sub):
                sub2 = os.path.join(sub, item2)
                if os.path.isdir(sub2) and item2 not in ('.', 'venv', '__pycache__'):
                    if os.path.exists(os.path.join(sub2, "app.py")):
                        return sub2
    return None

def sread(fp):
    try:
        with open(fp, 'r', encoding='utf-8', errors='replace') as f:
            return f.read()
    except:
        return ""

def get_code(root):
    c = {}
    c['app'] = sread(os.path.join(root, 'app.py'))
    tp = os.path.join(root, 'utils', 'text_tools.py')
    if not os.path.exists(tp):
        tp = os.path.join(root, 'text_tools.py')
    c['tt'] = sread(tp)
    vp = os.path.join(root, 'utils', 'validators.py')
    if not os.path.exists(vp):
        vp = os.path.join(root, 'validators.py')
    c['val'] = sread(vp)
    for fn in ['README.md', 'README.md.txt', 'README.md.txt.txt']:
        fp = os.path.join(root, fn)
        if os.path.exists(fp):
            c['readme'] = sread(fp)
            break
    else:
        c['readme'] = ""
    rp = os.path.join(root, 'outputs', 'results.txt')
    c['results'] = sread(rp) if os.path.exists(rp) else ""
    return c

def pfuncs(src):
    if not src.strip():
        return {}, False
    try:
        tree = ast.parse(src)
        return {n.name: n for n in ast.walk(tree) if isinstance(n, ast.FunctionDef)}, True
    except SyntaxError:
        return {}, False

def chk_struct(root, submission_path):
    s = {}
    s['app'] = os.path.exists(os.path.join(root, 'app.py'))
    s['utils'] = os.path.isdir(os.path.join(root, 'utils'))
    s['init'] = os.path.exists(os.path.join(root, 'utils', '__init__.py'))
    s['tt_utils'] = os.path.exists(os.path.join(root, 'utils', 'text_tools.py'))
    s['val_utils'] = os.path.exists(os.path.join(root, 'utils', 'validators.py'))
    s['tt_flat'] = os.path.exists(os.path.join(root, 'text_tools.py'))
    s['val_flat'] = os.path.exists(os.path.join(root, 'validators.py'))
    s['outdir'] = os.path.isdir(os.path.join(root, 'outputs'))
    s['restxt'] = os.path.exists(os.path.join(root, 'outputs', 'results.txt'))
    # Check for venv anywhere in the submission folder (search from top-level submission path)
    s['venv'] = False
    s['venv_loc'] = ""
    for dirpath, dirnames, filenames in os.walk(submission_path):
        # Skip already-found venv subdirs to avoid deep recursion
        dirnames[:] = [d for d in dirnames if d.lower() not in ('venv', '.venv', 'env', '.env', '__pycache__', 'Lib', 'Scripts', 'Include')]
        for d in list(os.listdir(dirpath)):
            if d.lower() in ('venv', '.venv', 'env', '.env'):
                candidate = os.path.join(dirpath, d)
                if os.path.isdir(candidate) and (
                    os.path.exists(os.path.join(candidate, 'pyvenv.cfg')) or
                    os.path.isdir(os.path.join(candidate, 'Scripts')) or
                    os.path.isdir(os.path.join(candidate, 'Lib'))):
                    s['venv'] = True
                    s['venv_loc'] = os.path.relpath(candidate, submission_path)
                    break
        if s['venv']:
            break
    s['readme'] = any(os.path.exists(os.path.join(root, f)) for f in ['README.md', 'README.md.txt', 'README.md.txt.txt'])
    s['fname'] = os.path.basename(root)
    return s

def ga(s, c):
    sc = 0; cm = []
    if s['app']:
        sc = 15; cm.append("app.py present")
    else:
        cm.append("MISSING app.py"); return 0, cm
    if s['venv']:
        sc -= 5; cm.append("DEDUCTION: venv/ included at '{}' (-5)".format(s['venv_loc']))
    return sc, cm

def gb(s, c):
    sc = 0; cm = []
    readme = c.get('readme', '')
    allcode = c['app'] + '\n' + c['tt'] + '\n' + c['val']
    if not readme.strip():
        cm.append("README missing or empty")
        for p in ['rich', 'regex', 'slugify', 'colorama', 'pyfiglet', 'tabulate']:
            if 'import ' + p in allcode or 'from ' + p in allcode:
                sc = 5; cm.append("Pkg '{}' in code but no README".format(p)); break
        return sc, cm
    rl = readme.lower()
    pf = None
    for p in ['rich', 'regex', 'python-slugify', 'slugify', 'colorama', 'pyfiglet', 'tabulate', 'termcolor']:
        if p in rl: pf = p; break
    if pf: sc += 5; cm.append("README mentions: " + pf)
    else: cm.append("README: no package name")
    if 'pip install' in rl: sc += 5; cm.append("README has install cmd")
    else: cm.append("README: no install cmd")
    rk = ['used for','used to','usage','rationale','reason','purpose','because','format','styling',
          'color','enhance','display','render','table','helps','allows','provides','enables',
          'terminal ui','better','clean','appealing','pretty','feature','output','styled','colored']
    if any(k in rl for k in rk): sc += 5; cm.append("README has rationale")
    elif len(readme.strip()) > 50: sc += 3; cm.append("README: weak rationale")
    else: sc += 1; cm.append("README: minimal")
    pu = False
    for p in ['rich','regex','slugify','colorama','pyfiglet','tabulate','termcolor']:
        if 'import ' + p in allcode or 'from ' + p in allcode: pu = True; break
    if not pu: sc = max(sc - 3, 0); cm.append("DEDUCTION: pkg not used in code (-3)")
    return min(sc, 15), cm

def gc(s, c):
    sc = 0.0; cm = []
    tt = c['tt']
    if not tt.strip(): cm.append("text_tools.py missing/empty"); return 0, cm
    sp = 0
    if not s['tt_utils'] and s['tt_flat']: sp = 2; cm.append("DEDUCTION: text_tools not in utils/ (-2)")
    fns, ok = pfuncs(tt)
    if not ok: cm.append("text_tools.py syntax errors"); return max(5-sp, 0), cm
    # clean_text
    cp = 0.0
    if 'clean_text' in fns:
        cp += 2
        if 'strip()' in tt: cp += 1.5
        if re.search(r're\.sub|split\(\).*join', tt): cp += 2
        if re.search(r'\[,\.!\?:;\]|punctuation', tt): cp += 2
        cp = min(cp, 7.5)
        cm.append("clean_text: {:.1f}/7.5".format(cp))
    else: cm.append("clean_text: MISSING")
    sc += cp
    # word_stats
    wp = 0.0
    hs = 'word_stats' in fns or 'get_stats' in fns
    if hs:
        if 'word_stats' in fns: wp += 1.5
        else: wp += 0.5; cm.append("  Note: get_stats not word_stats")
        m = 0
        if 'char_count' in tt or 'len(text)' in tt or 'Chars' in tt: m += 1
        if 'no_spaces' in tt or 'replace(" "' in tt or "replace(' '" in tt or 'no spaces' in tt.lower(): m += 1
        if 'word_count' in tt or 'Words' in tt or 'split()' in tt: m += 1
        if 'sentence' in tt.lower() or re.search(r'\[\.!\?\]', tt): m += 1
        if 'longest' in tt.lower() or 'max(' in tt: m += 1
        wp += min(m * 1.2, 6)
        wp = min(wp, 7.5)
        cm.append("word_stats: {:.1f}/7.5 [{}/5 metrics]".format(wp, m))
    else: cm.append("word_stats: MISSING")
    sc += wp
    # mask_sensitive
    mp = 0.0
    if 'mask_sensitive' in fns:
        mp += 1.5
        if re.search(r'@|email', tt.lower()): mp += 3
        if re.search(r'phone|\\d|digit|10', tt.lower()): mp += 3
        mp = min(mp, 7.5)
        cm.append("mask_sensitive: {:.1f}/7.5".format(mp))
    else: cm.append("mask_sensitive: MISSING")
    sc += mp
    # make_slug
    slp = 0.0
    if 'make_slug' in fns:
        slp += 1.5
        if 'lower()' in tt: slp += 1.5
        if 'replace' in tt or re.search(r'\\s.*-', tt): slp += 1.5
        if 'punctuation' in tt or re.search(r'\[\^a-z|re\.sub', tt): slp += 1.5
        if re.search(r'-\+|--', tt): slp += 1.5
        slp = min(slp, 7.5)
        cm.append("make_slug: {:.1f}/7.5".format(slp))
    else: cm.append("make_slug: MISSING")
    sc += slp
    sc = max(round(sc - sp, 1), 0)
    return min(sc, 30), cm

def gd(s, c):
    sc = 0.0; cm = []
    v = c['val']
    if not v.strip(): cm.append("validators.py missing/empty"); return 0, cm
    sp = 0
    if not s['val_utils'] and s['val_flat']: sp = 1; cm.append("DEDUCTION: validators not in utils/ (-1)")
    fns, ok = pfuncs(v)
    if not ok: cm.append("validators.py syntax errors"); return max(2-sp, 0), cm
    hr = 'require_non_empty' in fns or 'get_non_empty' in fns
    if hr:
        p = 0
        if 'require_non_empty' in fns: p += 1
        else: p += 0.5; cm.append("  Note: get_non_empty not require_non_empty")
        if 'input(' in v: p += 1.5
        if 'while' in v: p += 1.5
        if 'strip' in v: p += 1
        p = min(p, 5); sc += p
        cm.append("require_non_empty: {:.1f}/5".format(p))
    else: cm.append("require_non_empty: MISSING")
    hm = 'require_menu_choice' in fns or 'get_menu_choice' in fns
    if hm:
        p = 0
        if 'require_menu_choice' in fns: p += 1
        else: p += 0.5; cm.append("  Note: get_menu_choice not require_menu_choice")
        if 'input(' in v: p += 1.5
        if 'while' in v: p += 1.5
        if 'choices' in v or 'choice' in v or 'valid' in v: p += 1
        p = min(p, 5); sc += p
        cm.append("require_menu_choice: {:.1f}/5".format(p))
    else: cm.append("require_menu_choice: MISSING")
    sc = max(round(sc - sp, 1), 0)
    return min(sc, 10), cm

def ge(s, c):
    sc = 0.0; cm = []
    app = c['app']
    if not app.strip(): cm.append("app.py missing/empty"); return 0, cm
    _, ok = pfuncs(app)
    if not ok: cm.append("app.py syntax errors"); return 5, cm
    al = app.lower()
    # Menu
    mp = 0
    if 'while' in app and ('True' in app or 'true' in al): mp += 3
    elif 'while' in app: mp += 2
    mi = ['clean text','text statistics','mask sensitive','make slug','pipeline','full pipeline','process full','exit']
    fd = sum(1 for i in mi if i in al)
    if fd >= 5: mp += 3
    elif fd >= 3: mp += 2
    elif fd >= 1: mp += 1
    if re.search(r'print.*[1-6]|menu|option', al): mp += 2
    mp = min(mp, 8); sc += mp
    cm.append("Menu: {}/8".format(mp))
    # Inputs
    ip = 0
    if 'text_tools' in app or 'from utils' in app: ip += 2
    if 'validators' in app or 'require_non_empty' in app or 'get_non_empty' in app: ip += 2
    fc = ['clean_text','word_stats','get_stats','mask_sensitive','make_slug']
    cl = sum(1 for f in fc if f + '(' in app)
    if cl >= 4: ip += 4
    elif cl >= 3: ip += 3
    elif cl >= 2: ip += 2
    elif cl >= 1: ip += 1
    ip = min(ip, 8); sc += ip
    cm.append("Inputs: {}/8".format(ip))
    # Logging
    lp = 0
    hw = 'open(' in app and ('write' in app or 'append' in al)
    hf = 'log_to_file' in app or 'save_to_file' in app or 'log_result' in app
    if hw or hf: lp += 2
    if 'outputs' in app and 'results' in app: lp += 2
    elif 'results.txt' in app: lp += 1
    if 'datetime' in app or 'strftime' in app: lp += 2
    if 'operation' in al: lp += 1
    lp = min(lp, 7)
    if c['results'].strip(): cm.append("results.txt has logs")
    elif not s['restxt']: lp = max(lp - 1, 0); cm.append("results.txt missing/empty")
    sc += lp
    cm.append("Logging: {}/7".format(lp))
    # Package
    pp = 0; pks = []
    for p in ['rich','regex','slugify','colorama','pyfiglet','tabulate','termcolor']:
        if 'import ' + p in app or 'from ' + p in app: pks.append(p)
    if pks:
        pp += 3
        if 'rich' in pks:
            u = ['Console','Table','Panel','console.print','Style','Prompt']
            ud = sum(1 for x in u if x in app)
            if ud >= 3: pp += 4
            elif ud >= 2: pp += 3
            elif ud >= 1: pp += 2
        elif 'colorama' in pks:
            if 'Fore' in app or 'Style' in app: pp += 3
        elif 'pyfiglet' in pks:
            if 'figlet' in al: pp += 3
        else: pp += 2
        cm.append("Pkg: " + ', '.join(pks))
    else:
        for p in ['rich','regex','slugify','colorama','pyfiglet','tabulate']:
            if 'import ' + p in c['tt'] or 'from ' + p in c['tt']:
                pp += 4; cm.append("Pkg {} in text_tools only".format(p)); break
        if pp == 0: cm.append("No external pkg")
    pp = min(pp, 7); sc += pp
    cm.append("Pkg integration: {}/7".format(pp))
    return min(round(sc, 1), 30), cm

def gbon(c, root):
    sc = 0; cm = []
    app = c['app']; al = app.lower()
    ind = 0
    if re.search(r'["\']7["\']|7\.|choice.*7', al): ind += 1
    if re.search(r'file.*path|\.txt|enter.*file|load.*file|read.*file|from.*file', al): ind += 1
    if re.search(r'open\(.*["\']r["\']|\.read\(\)|readlines|read_from_file', app): ind += 1
    hs = any(os.path.exists(os.path.join(root, f)) for f in ['sample.txt','bonus.txt','test.txt'])
    if hs: ind += 1
    if ind >= 3: sc = 10; cm.append("File processing implemented")
    elif ind >= 2: sc = 7; cm.append("File processing partial")
    elif ind >= 1: sc = 3; cm.append("Minimal file code")
    return sc, cm

def grade(folder):
    path = os.path.join(BASE, folder)
    root = find_root(path)
    parts = folder.split('_')
    np = []
    for p in parts:
        if p.strip().isdigit(): break
        np.append(p.strip())
    name = ' '.join(np).replace('assignsubmission','').replace('file','').strip()
    if root is None:
        return {'name': name, 'a': 0, 'b': 0, 'c': 0, 'd': 0, 'e': 0, 'bon': 0, 'tot': 0, 'twb': 0, 'com': "NO SUBMISSION"}
    s = chk_struct(root, path)
    cd = get_code(root)
    asc, ac = ga(s, cd)
    bsc, bc = gb(s, cd)
    csc, cc = gc(s, cd)
    dsc, dc = gd(s, cd)
    esc, ec = ge(s, cd)
    bons, bonc = gbon(cd, root)
    tot = round(asc + bsc + csc + dsc + esc, 1)
    twb = round(tot + bons, 1)
    sn = []
    if s['fname'] != 'python_exam_strings': sn.append("Folder: " + s['fname'])
    if not s['utils'] and (s['tt_flat'] or s['val_flat']): sn.append("No utils/ dir")
    if s['utils'] and not s['init']: sn.append("No __init__.py")
    if not s['outdir']: sn.append("No outputs/ dir")
    if s['venv']: sn.append("venv/ INCLUDED at: " + s['venv_loc'])
    lines = []
    if sn: lines.append("STRUCTURE: " + "; ".join(sn))
    lines.append("Part A ({}/15): {}".format(asc, "; ".join(ac)))
    lines.append("Part B ({}/15): {}".format(bsc, "; ".join(bc)))
    lines.append("Part C ({}/30): {}".format(csc, "; ".join(cc)))
    lines.append("Part D ({}/10): {}".format(dsc, "; ".join(dc)))
    lines.append("Part E ({}/30): {}".format(esc, "; ".join(ec)))
    if bons > 0: lines.append("Bonus ({}/10): {}".format(bons, "; ".join(bonc)))
    return {'name': name, 'a': asc, 'b': bsc, 'c': csc, 'd': dsc, 'e': esc, 'bon': bons, 'tot': tot, 'twb': twb, 'com': "\n".join(lines)}

def make_report(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "Grade Summary"
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfl = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    sf = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    pf = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    wf = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ff = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bf = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('A1:L1')
    ws['A1'] = "IT 103 - Python Practical Exam Grade Report"
    ws['A1'].font = Font(bold=True, size=16, color="2F5496")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:L2')
    ws['A2'] = "Section: CAV-CCS-DCSN03C_2 | Generated: " + datetime.now().strftime('%B %d, %Y %I:%M %p')
    ws['A2'].font = Font(size=10, italic=True, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center')
    headers = ['#','Student Name','Part A\nVenv (15)','Part B\nPkg/README (15)','Part C\nStr Utils (30)',
               'Part D\nValidators (10)','Part E\nMain App (30)','Subtotal\n(/100)','Bonus\n(+10)',
               'Final\n(/110)','Pct','Remarks']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = hf; c.fill = hfl; c.border = tb
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[4].height = 55
    results.sort(key=lambda x: x['name'])
    for i, r in enumerate(results):
        row = i + 5
        ws.cell(row=row, column=1, value=i+1).border = tb
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        nc = ws.cell(row=row, column=2, value=r['name'])
        nc.border = tb; nc.font = Font(bold=True, size=10)
        for col, key in [(3,'a'),(4,'b'),(5,'c'),(6,'d'),(7,'e')]:
            c = ws.cell(row=row, column=col, value=r[key])
            c.border = tb; c.alignment = Alignment(horizontal='center'); c.number_format = '0.0'
        tc = ws.cell(row=row, column=8, value=r['tot'])
        tc.border = tb; tc.font = Font(bold=True, size=11); tc.alignment = Alignment(horizontal='center')
        tc.number_format = '0.0'
        if r['tot'] >= 75: tc.fill = pf
        elif r['tot'] >= 60: tc.fill = wf
        else: tc.fill = ff
        bc = ws.cell(row=row, column=9, value=r['bon'])
        bc.border = tb; bc.alignment = Alignment(horizontal='center')
        if r['bon'] > 0: bc.fill = bf
        fc = ws.cell(row=row, column=10, value=r['twb'])
        fc.border = tb; fc.font = Font(bold=True); fc.alignment = Alignment(horizontal='center')
        fc.number_format = '0.0'
        pc = ws.cell(row=row, column=11, value="{}%".format(round(r['tot'], 1)))
        pc.border = tb; pc.alignment = Alignment(horizontal='center')
        if r['tot'] >= 96: rem = "Outstanding"
        elif r['tot'] >= 90: rem = "Excellent"
        elif r['tot'] >= 80: rem = "Very Good"
        elif r['tot'] >= 75: rem = "Passed"
        elif r['tot'] >= 60: rem = "Needs Improvement"
        elif r['tot'] > 0: rem = "Failed"
        else: rem = "No Submission"
        rc = ws.cell(row=row, column=12, value=rem)
        rc.border = tb; rc.alignment = Alignment(horizontal='center')
        if rem in ('Outstanding','Excellent'): rc.font = Font(color="006100", bold=True)
        elif rem in ('Failed','No Submission'): rc.font = Font(color="9C0006", bold=True)
    sr = len(results) + 6
    ws.merge_cells('A{}:B{}'.format(sr, sr))
    ws.cell(row=sr, column=1, value="CLASS STATISTICS").font = Font(bold=True, size=12, color="2F5496")
    for c in range(1, 13): ws.cell(row=sr, column=c).fill = sf
    vt = [r['tot'] for r in results if r['tot'] > 0]
    stats = [("Total Students", len(results)), ("With Submissions", len(vt)),
             ("Class Average", "{:.1f}".format(sum(vt)/len(vt)) if vt else "N/A"),
             ("Highest Score", "{:.1f}".format(max(vt)) if vt else "N/A"),
             ("Lowest Score", "{:.1f}".format(min(vt)) if vt else "N/A"),
             ("Passing (>=75)", "{}/{}".format(sum(1 for t in vt if t >= 75), len(vt))),
             ("Failing (<75)", "{}/{}".format(sum(1 for t in vt if t < 75), len(vt))),
             ("With Bonus", "{}/{}".format(sum(1 for r in results if r['bon'] > 0), len(vt)))]
    for j, (label, val) in enumerate(stats):
        r = sr + 1 + j
        ws.merge_cells('A{}:B{}'.format(r, r))
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=3, value=str(val)).alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 32
    for l in ['C','D','E','F','G']: ws.column_dimensions[l].width = 14
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 18
    # Detailed sheet
    ws2 = wb.create_sheet("Detailed Feedback")
    ws2.merge_cells('A1:E1')
    ws2['A1'] = "Detailed Grading Feedback"
    ws2['A1'].font = Font(bold=True, size=14, color="2F5496")
    for col, h in enumerate(['#','Student Name','Score (/100)','Bonus (+10)','Detailed Comments'], 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = hf; c.fill = hfl; c.border = tb; c.alignment = Alignment(horizontal='center', vertical='center')
    for i, r in enumerate(results):
        row = i + 4
        ws2.cell(row=row, column=1, value=i+1).border = tb
        ws2.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=2, value=r['name']).border = tb
        ws2.cell(row=row, column=2).font = Font(bold=True)
        ws2.cell(row=row, column=3, value=r['tot']).border = tb
        ws2.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=3).number_format = '0.0'
        ws2.cell(row=row, column=4, value=r['bon']).border = tb
        ws2.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        cc = ws2.cell(row=row, column=5, value=r['com'])
        cc.border = tb; cc.alignment = Alignment(wrap_text=True, vertical='top')
        ws2.row_dimensions[row].height = 180
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 95
    out = os.path.join(BASE, "IT103_Practical_Exam_Grades.xlsx")
    wb.save(out)
    return out

def main():
    folders = sorted([i for i in os.listdir(BASE) if os.path.isdir(os.path.join(BASE, i)) and 'assignsubmission_file' in i])
    results = []
    output_lines = []
    output_lines.append("=" * 100)
    output_lines.append("IT 103 Python Practical Exam - Automated Grading | {} submissions".format(len(folders)))
    output_lines.append("=" * 100)
    output_lines.append("{:<35} {:>5} {:>5} {:>5} {:>5} {:>5} {:>4} {:>7}".format("Student","A/15","B/15","C/30","D/10","E/30","Bon","Total"))
    output_lines.append("-" * 100)
    for f in folders:
        r = grade(f)
        results.append(r)
        output_lines.append("{:<35} {:>5} {:>5} {:>5.1f} {:>5.1f} {:>5.1f} {:>4} {:>5.1f}/100".format(
            r['name'], r['a'], r['b'], r['c'], r['d'], r['e'], r['bon'], r['tot']))
    output_lines.append("=" * 100)
    path = make_report(results)
    vt = [r['tot'] for r in results if r['tot'] > 0]
    output_lines.append("")
    output_lines.append("CLASS SUMMARY")
    output_lines.append("-" * 40)
    output_lines.append("Total: {} students".format(len(results)))
    output_lines.append("Submissions: {}".format(len(vt)))
    if vt:
        output_lines.append("Average: {:.1f}/100".format(sum(vt)/len(vt)))
        output_lines.append("Highest: {:.1f}/100".format(max(vt)))
        output_lines.append("Lowest: {:.1f}/100".format(min(vt)))
        output_lines.append("Passing (>=75): {}/{}".format(sum(1 for t in vt if t >= 75), len(vt)))
    output_lines.append("")
    output_lines.append("Report saved: " + path)
    output_lines.append("=" * 100)
    # Write output to file
    out_file = os.path.join(BASE, "grading_output.txt")
    with open(out_file, 'w', encoding='utf-8') as f:
        f.write("\n".join(output_lines))
    # Also print
    for line in output_lines:
        print(line)

if __name__ == "__main__":
    main()
